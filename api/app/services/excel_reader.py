"""基幹システムが出力した Excel を読み、検証して正規化する。

DBには一切触れない。読み取りと検証だけを担い、結果を importer に渡す。
分離しておくと「投入せず検証だけ」のリクエストが素直に書ける。
"""

from __future__ import annotations

import datetime as dt
from dataclasses import dataclass, field
from decimal import Decimal
from typing import BinaryIO

from openpyxl import load_workbook

from app.domain import rules

# 基幹システムの出力列。順序も名称もこのとおりであることを検証する。
EXPECTED_COLUMNS: tuple[str, ...] = (
    "識別番号",
    "受注営業所コード",
    "受注営業所名",
    "販売先コード",
    "販売先名称",
    "販売先担当者",
    "営業担当者コード",
    "営業担当者名",
    "住所",
    "受注番号",
    "契約番号",
    "初回納品日",
    "引取日",
    "取引先発注番号1",
    "取引先発注番号2",
    "取引先発注番号3",
    "取引先発注番号4",
    "取引先発注番号5",
    "納品先名称",
    "納品日",
    "品番",
    "品名",
    "明細数量",
    "商品摘要",
    "レンタル開始日",
    "レンタル終了日",
    "レンタル期間",
    "請求期間開始日",
    "請求期間終了日",
    "経過日数",
    "月換算",
    "単位",
    "備考",
    "基本料",
    "月単価",
    "日単価",
    "販売単価",
    "レンタル単価",
    "レンタル金額",
    "販売金額",
    "小計",
    "合計",
    "二桁値引後計",
    "二桁値引",
    "表示順",
    "順番",
    "再発行",
    "削除",
    "暫定区分",
)


class Severity:
    ERROR = "ERROR"  # 投入不可
    WARNING = "WARNING"  # 投入は可能。人の確認を促す
    INFO = "INFO"


@dataclass
class Issue:
    severity: str
    type: str
    message: str
    rows: list[int] = field(default_factory=list)  # 該当するExcelの行番号


@dataclass
class SourceRow:
    """Excel1行ぶんを正規化したもの。"""

    excel_row: int

    # 組織・取引先
    office_code: str | None
    office_name: str | None
    customer_code: str | None
    customer_name: str | None
    customer_contact: str | None
    sales_rep_code: str | None
    sales_rep_name: str | None

    # 契約・現場
    contract_no: str
    site_name: str
    site_address: str | None
    client_name: str  # 納品先名称から名寄せした会社名

    # 受注
    order_no: str
    po_number_1: str | None
    po_number_2: str | None

    # 品目
    item_code: str
    item_name: str
    tax_category: str
    billing_group: str
    item_is_billable: bool

    # 明細
    delivery_date: dt.date | None
    return_date: dt.date | None
    shipped_date: dt.date | None
    rental_start: dt.date | None
    rental_end: dt.date | None
    unit: str | None
    note: str | None
    quantity: Decimal
    base_charge: Decimal | None
    unit_price: Decimal | None
    duration: Decimal | None
    unit_price_type: str
    display_order: int | None
    seq: int | None
    source_key: str | None
    is_provisional: bool
    is_deleted_in_source: bool

    # 突き合わせ用。アプリは金額を自前で計算するので投入はしない
    source_amount: Decimal | None


@dataclass
class ParseResult:
    rows: list[SourceRow]
    issues: list[Issue]
    period_start: dt.date | None
    period_end: dt.date | None
    customer_name: str | None

    @property
    def has_error(self) -> bool:
        return any(i.severity == Severity.ERROR for i in self.issues)

    @property
    def errors(self) -> list[Issue]:
        return [i for i in self.issues if i.severity == Severity.ERROR]

    @property
    def warnings(self) -> list[Issue]:
        return [i for i in self.issues if i.severity == Severity.WARNING]


def _to_date(value: object) -> dt.date | None:
    if value is None:
        return None
    if isinstance(value, dt.datetime):
        return value.date()
    if isinstance(value, dt.date):
        return value
    return None


def _to_int(value: object) -> int | None:
    d = rules.to_decimal(value)
    return int(d) if d is not None else None


def parse(source: BinaryIO | str) -> ParseResult:
    """Excelを読んで検証する。DBには触れない。

    エラーがある場合も、判明した範囲の情報は返す（画面に出すため）。
    """
    issues: list[Issue] = []

    wb = load_workbook(source, data_only=True, read_only=True)
    try:
        ws = wb.worksheets[0]
        grid = list(ws.iter_rows(values_only=True))
    finally:
        wb.close()

    if not grid:
        issues.append(
            Issue(Severity.ERROR, "EMPTY", "シートが空です。")
        )
        return ParseResult([], issues, None, None, None)

    header = tuple(
        (str(c).strip() if c is not None else "") for c in grid[0][: len(EXPECTED_COLUMNS)]
    )

    # ---- 列構成の検証 ---------------------------------------------------
    actual_width = len([c for c in grid[0] if c is not None and str(c).strip()])
    if header != EXPECTED_COLUMNS:
        missing = [c for c in EXPECTED_COLUMNS if c not in header]
        extra = [c for c in header if c and c not in EXPECTED_COLUMNS]
        detail = []
        if actual_width != len(EXPECTED_COLUMNS):
            detail.append(f"列数が {actual_width}（期待 {len(EXPECTED_COLUMNS)}）")
        if missing:
            detail.append(f"不足: {'、'.join(missing[:5])}")
        if extra:
            detail.append(f"想定外: {'、'.join(extra[:5])}")
        if not detail:
            detail.append("列の並び順が異なります")
        issues.append(
            Issue(
                Severity.ERROR,
                "FORMAT",
                "列構成が想定と異なります。" + " / ".join(detail),
            )
        )
        return ParseResult([], issues, None, None, None)

    col = {name: i for i, name in enumerate(EXPECTED_COLUMNS)}

    # ---- 明細行の読み取り -----------------------------------------------
    rows: list[SourceRow] = []
    periods: set[tuple[dt.date | None, dt.date | None]] = set()
    customers: set[str] = set()
    zero_price_rows: list[int] = []
    missing_key_rows: list[int] = []

    def cell(raw: tuple, name: str) -> object:
        idx = col[name]
        return raw[idx] if idx < len(raw) else None

    for excel_row, raw in enumerate(grid[1:], start=2):
        if raw is None or all(v is None or str(v).strip() == "" for v in raw):
            continue  # 空行は読み飛ばす

        contract_no = rules.to_text(cell(raw, "契約番号"))
        site_name = rules.to_text(cell(raw, "納品先名称"))
        item_code = rules.to_text(cell(raw, "品番"))
        order_no = rules.to_text(cell(raw, "受注番号"))

        if not (contract_no and site_name and item_code and order_no):
            missing_key_rows.append(excel_row)
            continue

        item_name = rules.to_text(cell(raw, "品名")) or ""

        unit_price_type, unit_price, duration = rules.resolve_unit_price(
            monthly_rate=rules.to_decimal(cell(raw, "月単価")),
            rental_rate=rules.to_decimal(cell(raw, "レンタル単価")),
            monthly_conversion=rules.to_decimal(cell(raw, "月換算")),
            elapsed_days=rules.to_decimal(cell(raw, "経過日数")),
            sale_price=rules.to_decimal(cell(raw, "販売単価")),
        )

        if unit_price is None or unit_price == 0:
            zero_price_rows.append(excel_row)

        quantity = rules.to_decimal(cell(raw, "明細数量")) or Decimal(0)

        rows.append(
            SourceRow(
                excel_row=excel_row,
                office_code=rules.to_text(cell(raw, "受注営業所コード")),
                office_name=rules.to_text(cell(raw, "受注営業所名")),
                customer_code=rules.to_text(cell(raw, "販売先コード")),
                customer_name=rules.to_text(cell(raw, "販売先名称")),
                customer_contact=rules.to_text(cell(raw, "販売先担当者")),
                sales_rep_code=rules.to_text(cell(raw, "営業担当者コード")),
                sales_rep_name=rules.to_text(cell(raw, "営業担当者名")),
                contract_no=contract_no,
                site_name=site_name,
                site_address=rules.to_text(cell(raw, "住所")),
                client_name=rules.normalize_client_name(site_name),
                order_no=order_no,
                po_number_1=rules.to_text(cell(raw, "取引先発注番号1")),
                po_number_2=rules.to_text(cell(raw, "取引先発注番号2")),
                item_code=item_code,
                item_name=item_name,
                tax_category=rules.classify_tax_category(item_code),
                billing_group=rules.classify_billing_group(item_code),
                item_is_billable=rules.is_billable_item(item_name),
                delivery_date=_to_date(cell(raw, "初回納品日")),
                return_date=_to_date(cell(raw, "引取日")),
                shipped_date=_to_date(cell(raw, "納品日")),
                rental_start=_to_date(cell(raw, "レンタル開始日")),
                rental_end=_to_date(cell(raw, "レンタル終了日")),
                unit=rules.to_text(cell(raw, "単位")),
                note=rules.to_text(cell(raw, "商品摘要")),
                quantity=quantity,
                base_charge=rules.normalize_base_charge(
                    rules.to_decimal(cell(raw, "基本料"))
                ),
                unit_price=unit_price,
                duration=duration,
                unit_price_type=unit_price_type,
                display_order=_to_int(cell(raw, "表示順")),
                seq=_to_int(cell(raw, "順番")),
                source_key=rules.to_text(cell(raw, "識別番号")),
                is_provisional=bool(rules.to_text(cell(raw, "暫定区分"))),
                # 基幹システム側で取消・削除された行。値の有無だけを見る
                # （VBAはこの列を一切参照していないため判定記号の前例がない。
                #  空でなければ削除扱いとする）
                is_deleted_in_source=bool(rules.to_text(cell(raw, "削除"))),
                source_amount=rules.to_decimal(cell(raw, "小計")),
            )
        )

        periods.add(
            (
                _to_date(cell(raw, "請求期間開始日")),
                _to_date(cell(raw, "請求期間終了日")),
            )
        )
        cust = rules.to_text(cell(raw, "販売先名称"))
        if cust:
            customers.add(cust)

    if not rows:
        issues.append(Issue(Severity.ERROR, "EMPTY", "明細行が1行もありません。"))
        return ParseResult([], issues, None, None, None)

    if missing_key_rows:
        issues.append(
            Issue(
                Severity.WARNING,
                "MISSING_KEY",
                f"契約番号・受注番号・納品先名称・品番のいずれかが空の行を "
                f"{len(missing_key_rows)}行スキップしました。",
                missing_key_rows[:20],
            )
        )

    # ---- 請求期間の単一性 -----------------------------------------------
    period_start = period_end = None
    valid_periods = {p for p in periods if p[0] and p[1]}
    if not valid_periods:
        issues.append(
            Issue(
                Severity.ERROR,
                "NO_PERIOD",
                "請求期間開始日・終了日が読み取れませんでした。",
            )
        )
    elif len(valid_periods) > 1:
        shown = "、".join(
            f"{s:%Y-%m-%d}〜{e:%Y-%m-%d}" for s, e in sorted(valid_periods)[:4]
        )
        issues.append(
            Issue(
                Severity.ERROR,
                "MULTI_PERIOD",
                f"請求期間が {len(valid_periods)} 種類混在しています（{shown}）。"
                "基幹システム側の絞り込み条件を確認してください。",
            )
        )
    else:
        period_start, period_end = next(iter(valid_periods))

    # ---- 販売先の単一性 -------------------------------------------------
    customer_name = None
    if len(customers) > 1:
        issues.append(
            Issue(
                Severity.ERROR,
                "MULTI_CUSTOMER",
                f"販売先が {len(customers)} 社混在しています"
                f"（{'、'.join(sorted(customers)[:3])}）。"
                "1ファイル＝1販売先で出力してください。",
            )
        )
    elif customers:
        customer_name = next(iter(customers))

    # ---- 単価0の警告 -----------------------------------------------------
    if zero_price_rows:
        issues.append(
            Issue(
                Severity.WARNING,
                "ZERO_PRICE",
                f"単価が0または空の行が {len(zero_price_rows)} 行あります。",
                zero_price_rows[:20],
            )
        )

    return ParseResult(rows, issues, period_start, period_end, customer_name)
